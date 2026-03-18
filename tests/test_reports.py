from __future__ import annotations

import io
from datetime import UTC, datetime
from types import SimpleNamespace
from typing import Any, cast
from unittest.mock import AsyncMock
from uuid import uuid4

import pytest
from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from redis.asyncio import Redis
from sqlalchemy.ext.asyncio import AsyncSession

from app.auth.dependencies import get_current_user
from app.main import app
from app.models.enums import UserRoleEnum
from app.routes import analytics as analytics_routes
from app.schemas.analytics import (
    AlertsResponse,
    EnsembleYieldForecastResponse,
    FarmStatusResponse,
    IrrigationScheduleResponse,
    NutrientReportResponse,
    ZoneAlerts,
    ZoneStatus,
)
from app.schemas.reports import ReportRequest
from app.services.analytics_service import AnalyticsService
from app.services.farm_service import FarmService
from app.services.report_service import ReportService


def _build_report_bytes() -> io.BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Farm Summary"
    sheet.append(["ok"])
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _fake_farm() -> Any:
    farm_id = uuid4()
    zone = SimpleNamespace(
        id=uuid4(),
        name="Zone A",
        zone_type=SimpleNamespace(value="greenhouse"),
        area_m2=120.0,
        soil_type="loam",
    )
    return SimpleNamespace(
        id=farm_id,
        name="Demo Farm",
        farm_type=SimpleNamespace(value="greenhouse"),
        timezone="UTC",
        zones=[zone],
    )


def _fake_farm_status(zone_id: Any) -> FarmStatusResponse:
    return FarmStatusResponse(
        farm_id=uuid4(),
        generated_at=datetime.now(UTC),
        zones=[
            ZoneStatus(
                zone_id=zone_id,
                query_vertex_id=uuid4(),
                status={"soil": {"ok": True}, "weather": {"ok": True}},
            )
        ],
    )


def _fake_irrigation(farm_id: Any) -> IrrigationScheduleResponse:
    return IrrigationScheduleResponse(
        farm_id=farm_id,
        horizon_days=7,
        cached=False,
        generated_at=datetime.now(UTC),
        items=[
            {
                "zone_id": "zone-1",
                "day": "2026-03-18",
                "irrigate": True,
                "volume_liters": 32.0,
                "priority": 0.84,
                "projected_moisture": 0.31,
                "trigger_reason": "Moisture below threshold",
            }
        ],
    )


def _fake_yield(farm_id: Any) -> EnsembleYieldForecastResponse:
    return EnsembleYieldForecastResponse(
        farm_id=farm_id,
        generated_at=datetime.now(UTC),
        include_members=True,
        ensemble_weights={
            "fao_single": 0.3,
            "exp_smoothing": 0.4,
            "quantile_regression": 0.3,
        },
        items=[
            {
                "crop_bed_id": "bed-1",
                "yield_estimate_kg_m2": 4.2,
                "yield_lower": 3.8,
                "yield_upper": 4.9,
                "confidence": 0.86,
                "stress_factors": {"Ks": 0.8, "Kn": 0.9, "Kl": 0.85, "Kw": 0.88},
                "model_layer": "ensemble",
                "ensemble_weights": {
                    "fao_single": 0.3,
                    "exp_smoothing": 0.4,
                    "quantile_regression": 0.3,
                },
                "hyperparameters": {"exp_alpha": 0.3},
                "ensemble_members": [
                    {
                        "model_name": "fao_single",
                        "yield_estimate": 4.1,
                        "lower": 3.7,
                        "upper": 4.8,
                        "weight": 0.3,
                    }
                ],
            }
        ],
    )


def _fake_nutrients(farm_id: Any) -> NutrientReportResponse:
    return NutrientReportResponse(
        farm_id=farm_id,
        generated_at=datetime.now(UTC),
        items=[
            {
                "zone_id": "zone-1",
                "nitrogen_deficit": 0.2,
                "phosphorus_deficit": 0.1,
                "potassium_deficit": 0.15,
                "urgency": "high",
                "suggested_amendment": "Add balanced NPK",
                "visual_confirmed": True,
            }
        ],
    )


def _fake_alerts(farm_id: Any) -> AlertsResponse:
    return AlertsResponse(
        farm_id=farm_id,
        generated_at=datetime.now(UTC),
        zones=[
            ZoneAlerts(
                zone_id=uuid4(),
                alerts=[
                    {
                        "source": "anomaly",
                        "severity": "warning",
                        "payload": {
                            "layer": "vision",
                            "anomaly_type": "wilting",
                            "urgency": "medium",
                            "suggestion": "Inspect affected crop bed",
                        },
                    }
                ],
            )
        ],
    )


def _fake_graph() -> dict[str, Any]:
    return {
        "farm_id": "farm-1",
        "layers": {
            "soil": {
                "history_length": 4,
                "history_head": 1,
                "feature_history": [
                    [
                        [0.30, 0.32, 0.33, 0.31],
                        [22.0, 22.1, 22.2, 22.3],
                        [1.2, 1.1, 1.2, 1.3],
                        [6.5, 6.4, 6.5, 6.6],
                    ]
                ],
            }
        },
    }


@pytest.mark.asyncio
async def test_report_endpoint_returns_attachment_headers(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    async def fake_generate_artifact(
        self: ReportService,
        farm_id: Any,
        payload: ReportRequest,
        *,
        output_format: str,
    ) -> tuple[bytes, str, str, bool]:
        assert payload.irrigation_horizon_days == 7
        assert output_format == "xlsx"
        return (
            _build_report_bytes().getvalue(),
            "agrisense-report-test.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            False,
        )

    monkeypatch.setattr(ReportService, "generate_report_artifact", fake_generate_artifact)

    response = await client.post(
        f"/api/v1/analytics/{uuid4()}/reports/generate",
        json={
            "irrigation_horizon_days": 7,
            "include_members": True,
            "include_history_charts": True,
        },
    )
    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert (
        'attachment; filename="agrisense-report-test.xlsx"'
        in response.headers["content-disposition"]
    )
    assert response.headers["x-report-cache"] == "MISS"


@pytest.mark.asyncio
async def test_report_endpoint_pdf_response(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    async def fake_generate_artifact(
        self: ReportService,
        farm_id: Any,
        payload: ReportRequest,
        *,
        output_format: str,
    ) -> tuple[bytes, str, str, bool]:
        assert output_format == "pdf"
        return (b"%PDF-1.4\n%...", "agrisense-report-test.pdf", "application/pdf", True)

    monkeypatch.setattr(ReportService, "generate_report_artifact", fake_generate_artifact)

    response = await client.post(
        f"/api/v1/analytics/{uuid4()}/reports/generate?format=pdf",
        json={
            "irrigation_horizon_days": 7,
            "include_members": True,
            "include_history_charts": True,
        },
    )
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/pdf"
    assert (
        'attachment; filename="agrisense-report-test.pdf"'
        in response.headers["content-disposition"]
    )
    assert response.headers["x-report-cache"] == "HIT"


@pytest.mark.asyncio
async def test_report_service_builds_valid_workbook(monkeypatch: pytest.MonkeyPatch) -> None:
    farm = _fake_farm()

    async def fake_get_farm(self: FarmService, farm_id: Any) -> Any:
        return farm

    async def fake_get_graph(self: FarmService, farm_id: Any) -> dict[str, Any]:
        return _fake_graph()

    async def fake_farm_status(self: AnalyticsService, farm_id: Any) -> FarmStatusResponse:
        return _fake_farm_status(farm.zones[0].id)

    async def fake_irrigation(self: AnalyticsService, farm_id: Any, horizon_days: int) -> Any:
        return _fake_irrigation(farm.id)

    async def fake_yield(
        self: AnalyticsService,
        farm_id: Any,
        include_members: bool,
    ) -> Any:
        return _fake_yield(farm.id)

    async def fake_nutrients(self: AnalyticsService, farm_id: Any) -> Any:
        return _fake_nutrients(farm.id)

    async def fake_alerts(self: AnalyticsService, farm_id: Any) -> Any:
        return _fake_alerts(farm.id)

    monkeypatch.setattr(FarmService, "get_farm", fake_get_farm)
    monkeypatch.setattr(FarmService, "get_graph", fake_get_graph)
    monkeypatch.setattr(AnalyticsService, "get_farm_status", fake_farm_status)
    monkeypatch.setattr(AnalyticsService, "get_irrigation_schedule", fake_irrigation)
    monkeypatch.setattr(AnalyticsService, "get_ensemble_yield_forecast", fake_yield)
    monkeypatch.setattr(AnalyticsService, "get_nutrient_report", fake_nutrients)
    monkeypatch.setattr(AnalyticsService, "get_active_alerts", fake_alerts)

    service = ReportService(cast(AsyncSession, object()), cast(Redis, None))
    report_bytes = await service.generate(
        uuid4(),
        ReportRequest(
            irrigation_horizon_days=7,
            include_members=True,
            include_history_charts=True,
        ),
    )

    workbook = load_workbook(io.BytesIO(report_bytes.getvalue()))
    assert workbook.sheetnames == [
        "Farm Summary",
        "Irrigation Schedule",
        "Yield Forecast",
        "Nutrient Status",
        "Active Alerts",
        "Feature History",
    ]

    summary_sheet = workbook["Farm Summary"]
    assert summary_sheet["A8"].value == "Zone ID"
    assert summary_sheet["B8"].value == "Zone Name"

    irrigation_sheet = workbook["Irrigation Schedule"]
    assert irrigation_sheet["A1"].value == "Zone"
    assert irrigation_sheet["E1"].value == "Priority"
    irrigation_rule_count = sum(
        len(entry)
        for entry in irrigation_sheet.conditional_formatting._cf_rules.values()  # type: ignore[attr-defined]
    )
    assert irrigation_rule_count >= 3

    yield_sheet = workbook["Yield Forecast"]
    weight_summary_labels = {
        str(cell.value)
        for row in yield_sheet.iter_rows(min_col=1, max_col=2)
        for cell in row
        if cell.value is not None
    }
    assert "Ensemble Weights Summary" in weight_summary_labels

    history_sheet = workbook["Feature History"]
    assert history_sheet["B2"].value == 0
    assert history_sheet["B3"].value == 15


@pytest.mark.asyncio
async def test_report_service_sync_cache_hits_on_repeat(monkeypatch: pytest.MonkeyPatch) -> None:
    class InMemoryRedis:
        def __init__(self) -> None:
            self._store: dict[str, bytes] = {}

        async def get(self, key: str) -> bytes | None:
            return self._store.get(key)

        async def setex(self, key: str, ttl: int, value: bytes) -> bool:
            self._store[key] = value
            return True

    calls = {"count": 0}

    async def fake_generate(
        self: ReportService,
        farm_id: Any,
        payload: ReportRequest,
    ) -> io.BytesIO:
        calls["count"] += 1
        return _build_report_bytes()

    monkeypatch.setattr(ReportService, "generate", fake_generate)

    service = ReportService(cast(AsyncSession, object()), cast(Redis, InMemoryRedis()))
    req = ReportRequest(include_members=False, include_history_charts=False)
    farm_id = uuid4()

    content_one, _, media_type_one, cache_one = await service.generate_report_artifact(
        farm_id, req, output_format="xlsx"
    )
    content_two, _, media_type_two, cache_two = await service.generate_report_artifact(
        farm_id, req, output_format="xlsx"
    )

    assert media_type_one == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert media_type_two == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert cache_one is False
    assert cache_two is True
    assert content_one == content_two
    assert calls["count"] == 1


@pytest.mark.asyncio
async def test_yield_sheet_contains_line_chart(monkeypatch: pytest.MonkeyPatch) -> None:
    farm = _fake_farm()

    async def fake_get_farm(self: FarmService, farm_id: Any) -> Any:
        return farm

    async def fake_farm_status(self: AnalyticsService, farm_id: Any) -> Any:
        return _fake_farm_status(farm.zones[0].id)

    async def fake_irrigation(self: AnalyticsService, farm_id: Any, horizon_days: int) -> Any:
        return _fake_irrigation(farm.id)

    async def fake_yield(self: AnalyticsService, farm_id: Any, include_members: bool) -> Any:
        return _fake_yield(farm.id)

    async def fake_nutrients(self: AnalyticsService, farm_id: Any) -> Any:
        return _fake_nutrients(farm.id)

    async def fake_alerts(self: AnalyticsService, farm_id: Any) -> Any:
        return _fake_alerts(farm.id)

    monkeypatch.setattr(FarmService, "get_farm", fake_get_farm)
    monkeypatch.setattr(AnalyticsService, "get_farm_status", fake_farm_status)
    monkeypatch.setattr(AnalyticsService, "get_irrigation_schedule", fake_irrigation)
    monkeypatch.setattr(AnalyticsService, "get_ensemble_yield_forecast", fake_yield)
    monkeypatch.setattr(AnalyticsService, "get_nutrient_report", fake_nutrients)
    monkeypatch.setattr(AnalyticsService, "get_active_alerts", fake_alerts)

    service = ReportService(cast(AsyncSession, object()), cast(Redis, None))
    report_bytes = await service.generate(
        uuid4(),
        ReportRequest(include_members=True, include_history_charts=False),
    )

    workbook = load_workbook(io.BytesIO(report_bytes.getvalue()))
    ws = workbook["Yield Forecast"]
    assert len(ws._charts) >= 1


@pytest.mark.asyncio
async def test_nutrient_sheet_has_conditional_formatting(monkeypatch: pytest.MonkeyPatch) -> None:
    farm = _fake_farm()

    async def fake_get_farm(self: FarmService, farm_id: Any) -> Any:
        return farm

    async def fake_farm_status(self: AnalyticsService, farm_id: Any) -> Any:
        return _fake_farm_status(farm.zones[0].id)

    async def fake_irrigation(self: AnalyticsService, farm_id: Any, horizon_days: int) -> Any:
        return _fake_irrigation(farm.id)

    async def fake_yield(self: AnalyticsService, farm_id: Any, include_members: bool) -> Any:
        return _fake_yield(farm.id)

    async def fake_nutrients(self: AnalyticsService, farm_id: Any) -> Any:
        return _fake_nutrients(farm.id)

    async def fake_alerts(self: AnalyticsService, farm_id: Any) -> Any:
        return _fake_alerts(farm.id)

    monkeypatch.setattr(FarmService, "get_farm", fake_get_farm)
    monkeypatch.setattr(AnalyticsService, "get_farm_status", fake_farm_status)
    monkeypatch.setattr(AnalyticsService, "get_irrigation_schedule", fake_irrigation)
    monkeypatch.setattr(AnalyticsService, "get_ensemble_yield_forecast", fake_yield)
    monkeypatch.setattr(AnalyticsService, "get_nutrient_report", fake_nutrients)
    monkeypatch.setattr(AnalyticsService, "get_active_alerts", fake_alerts)

    service = ReportService(cast(AsyncSession, object()), cast(Redis, None))
    report_bytes = await service.generate(uuid4(), ReportRequest(include_history_charts=False))

    workbook = load_workbook(io.BytesIO(report_bytes.getvalue()))
    ws = workbook["Nutrient Status"]
    assert len(ws.conditional_formatting) > 0


@pytest.mark.asyncio
async def test_alerts_sheet_enables_autofilter(monkeypatch: pytest.MonkeyPatch) -> None:
    farm = _fake_farm()

    async def fake_get_farm(self: FarmService, farm_id: Any) -> Any:
        return farm

    async def fake_farm_status(self: AnalyticsService, farm_id: Any) -> Any:
        return _fake_farm_status(farm.zones[0].id)

    async def fake_irrigation(self: AnalyticsService, farm_id: Any, horizon_days: int) -> Any:
        return _fake_irrigation(farm.id)

    async def fake_yield(self: AnalyticsService, farm_id: Any, include_members: bool) -> Any:
        return _fake_yield(farm.id)

    async def fake_nutrients(self: AnalyticsService, farm_id: Any) -> Any:
        return _fake_nutrients(farm.id)

    async def fake_alerts(self: AnalyticsService, farm_id: Any) -> Any:
        return _fake_alerts(farm.id)

    monkeypatch.setattr(FarmService, "get_farm", fake_get_farm)
    monkeypatch.setattr(AnalyticsService, "get_farm_status", fake_farm_status)
    monkeypatch.setattr(AnalyticsService, "get_irrigation_schedule", fake_irrigation)
    monkeypatch.setattr(AnalyticsService, "get_ensemble_yield_forecast", fake_yield)
    monkeypatch.setattr(AnalyticsService, "get_nutrient_report", fake_nutrients)
    monkeypatch.setattr(AnalyticsService, "get_active_alerts", fake_alerts)

    service = ReportService(cast(AsyncSession, object()), cast(Redis, None))
    report_bytes = await service.generate(uuid4(), ReportRequest(include_history_charts=False))

    workbook = load_workbook(io.BytesIO(report_bytes.getvalue()))
    ws = workbook["Active Alerts"]
    assert ws.auto_filter.ref is not None


@pytest.mark.asyncio
async def test_history_sheet_presence_toggled(monkeypatch: pytest.MonkeyPatch) -> None:
    farm = _fake_farm()

    async def fake_get_farm(self: FarmService, farm_id: Any) -> Any:
        return farm

    async def fake_get_graph(self: FarmService, farm_id: Any) -> dict[str, Any]:
        return _fake_graph()

    async def fake_farm_status(self: AnalyticsService, farm_id: Any) -> Any:
        return _fake_farm_status(farm.zones[0].id)

    async def fake_irrigation(self: AnalyticsService, farm_id: Any, horizon_days: int) -> Any:
        return _fake_irrigation(farm.id)

    async def fake_yield(self: AnalyticsService, farm_id: Any, include_members: bool) -> Any:
        return _fake_yield(farm.id)

    async def fake_nutrients(self: AnalyticsService, farm_id: Any) -> Any:
        return _fake_nutrients(farm.id)

    async def fake_alerts(self: AnalyticsService, farm_id: Any) -> Any:
        return _fake_alerts(farm.id)

    monkeypatch.setattr(FarmService, "get_farm", fake_get_farm)
    monkeypatch.setattr(FarmService, "get_graph", fake_get_graph)
    monkeypatch.setattr(AnalyticsService, "get_farm_status", fake_farm_status)
    monkeypatch.setattr(AnalyticsService, "get_irrigation_schedule", fake_irrigation)
    monkeypatch.setattr(AnalyticsService, "get_ensemble_yield_forecast", fake_yield)
    monkeypatch.setattr(AnalyticsService, "get_nutrient_report", fake_nutrients)
    monkeypatch.setattr(AnalyticsService, "get_active_alerts", fake_alerts)

    service = ReportService(cast(AsyncSession, object()), cast(Redis, None))

    with_history = await service.generate(uuid4(), ReportRequest(include_history_charts=True))
    wb_with_history = load_workbook(io.BytesIO(with_history.getvalue()))
    assert "Feature History" in wb_with_history.sheetnames

    without_history = await service.generate(uuid4(), ReportRequest(include_history_charts=False))
    wb_without_history = load_workbook(io.BytesIO(without_history.getvalue()))
    assert "Feature History" not in wb_without_history.sheetnames


@pytest.mark.asyncio
async def test_report_endpoint_maps_missing_farm_to_404(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    async def fake_generate_artifact(
        self: ReportService,
        farm_id: Any,
        payload: ReportRequest,
        *,
        output_format: str,
    ) -> tuple[bytes, str, str, bool]:
        raise LookupError("Farm not found")

    monkeypatch.setattr(ReportService, "generate_report_artifact", fake_generate_artifact)

    response = await client.post(
        f"/api/v1/analytics/{uuid4()}/reports/generate",
        json={
            "irrigation_horizon_days": 7,
            "include_members": True,
            "include_history_charts": True,
        },
    )
    assert response.status_code == 404


@pytest.mark.asyncio
async def test_report_endpoint_openapi_contract(client: AsyncClient) -> None:
    response = await client.get("/openapi.json")
    assert response.status_code == 200
    paths = response.json()["paths"]
    assert "/api/v1/analytics/{farm_id}/reports/generate" in paths
    assert "/api/v1/analytics/{farm_id}/reports/generate/async" in paths
    assert "/api/v1/analytics/{farm_id}/reports/jobs/{job_id}" in paths
    assert "/api/v1/analytics/{farm_id}/reports/jobs/{job_id}/download" in paths


@pytest.mark.asyncio
async def test_report_async_enqueue_returns_202(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    farm_id = uuid4()
    job_id = uuid4()
    created_at = datetime.now(UTC)

    async def fake_create(
        self: ReportService,
        _farm_id: Any,
        payload: ReportRequest,
        *,
        output_format: str,
    ) -> Any:
        assert payload.include_members is True
        assert output_format == "xlsx"
        return SimpleNamespace(
            job_id=job_id,
            farm_id=farm_id,
            status="queued",
            created_at=created_at,
        )

    async def fake_run_job(
        job_id: Any,
        farm_id: Any,
        payload: ReportRequest,
        output_format: str,
        redis_client: Any,
    ) -> None:
        assert output_format == "xlsx"
        return None

    monkeypatch.setattr(ReportService, "create_report_job", fake_create)
    monkeypatch.setattr(analytics_routes, "_run_report_job", fake_run_job)
    app.state.redis = SimpleNamespace(
        incr=AsyncMock(return_value=1),
        expire=AsyncMock(return_value=True),
        setex=AsyncMock(return_value=True),
    )

    response = await client.post(
        f"/api/v1/analytics/{farm_id}/reports/generate/async",
        json={
            "irrigation_horizon_days": 7,
            "include_members": True,
            "include_history_charts": True,
        },
    )
    assert response.status_code == 202
    body = response.json()
    assert body["job_id"] == str(job_id)
    assert body["status"] == "queued"


@pytest.mark.asyncio
async def test_report_async_status_endpoint(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    farm_id = uuid4()
    job_id = uuid4()
    now = datetime.now(UTC)

    async def fake_status(self: ReportService, _job_id: Any) -> Any:
        return SimpleNamespace(
            job_id=job_id,
            farm_id=farm_id,
            status="succeeded",
            created_at=now,
            updated_at=now,
            completed_at=now,
            error=None,
            filename="report.xlsx",
            details={"size_bytes": 128},
        )

    monkeypatch.setattr(ReportService, "get_report_job_status", fake_status)
    app.state.redis = None

    response = await client.get(f"/api/v1/analytics/{farm_id}/reports/jobs/{job_id}")
    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "succeeded"
    assert body["filename"] == "report.xlsx"


@pytest.mark.asyncio
async def test_report_async_download_endpoint(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    farm_id = uuid4()
    job_id = uuid4()
    now = datetime.now(UTC)

    async def fake_status(self: ReportService, _job_id: Any) -> Any:
        return SimpleNamespace(
            job_id=job_id,
            farm_id=farm_id,
            status="succeeded",
            created_at=now,
            updated_at=now,
            completed_at=now,
            error=None,
            filename="report.xlsx",
            details={},
        )

    async def fake_file(self: ReportService, _job_id: Any) -> tuple[bytes, str, str]:
        return (
            b"xlsx-bytes",
            "report.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    monkeypatch.setattr(ReportService, "get_report_job_status", fake_status)
    monkeypatch.setattr(ReportService, "get_report_job_file", fake_file)
    app.state.redis = None

    response = await client.get(f"/api/v1/analytics/{farm_id}/reports/jobs/{job_id}/download")
    assert response.status_code == 200
    assert response.content == b"xlsx-bytes"
    assert 'attachment; filename="report.xlsx"' in response.headers["content-disposition"]


@pytest.mark.asyncio
async def test_report_async_download_not_ready_returns_409(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    farm_id = uuid4()
    job_id = uuid4()
    now = datetime.now(UTC)

    async def fake_status(self: ReportService, _job_id: Any) -> Any:
        return SimpleNamespace(
            job_id=job_id,
            farm_id=farm_id,
            status="running",
            created_at=now,
            updated_at=now,
            completed_at=None,
            error=None,
            filename=None,
            details={},
        )

    async def fake_file(self: ReportService, _job_id: Any) -> tuple[bytes, str, str]:
        raise ValueError("report job not ready")

    monkeypatch.setattr(ReportService, "get_report_job_status", fake_status)
    monkeypatch.setattr(ReportService, "get_report_job_file", fake_file)
    app.state.redis = None

    response = await client.get(f"/api/v1/analytics/{farm_id}/reports/jobs/{job_id}/download")
    assert response.status_code == 409


@pytest.mark.asyncio
async def test_report_endpoint_rbac_readonly_forbidden(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    async def fake_generate(
        self: ReportService, farm_id: Any, payload: ReportRequest
    ) -> io.BytesIO:
        return _build_report_bytes()

    async def readonly_user() -> Any:
        return SimpleNamespace(
            id=uuid4(),
            role=UserRoleEnum.readonly,
            is_active=True,
            email="readonly@test.local",
        )

    monkeypatch.setattr(ReportService, "generate", fake_generate)
    app.dependency_overrides[get_current_user] = readonly_user
    try:
        response = await client.post(
            f"/api/v1/analytics/{uuid4()}/reports/generate",
            json={
                "irrigation_horizon_days": 7,
                "include_members": True,
                "include_history_charts": True,
            },
        )
        assert response.status_code == 403
    finally:
        app.dependency_overrides.pop(get_current_user, None)


@pytest.mark.asyncio
async def test_report_endpoint_rbac_field_operator_forbidden(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    async def fake_generate(
        self: ReportService, farm_id: Any, payload: ReportRequest
    ) -> io.BytesIO:
        return _build_report_bytes()

    async def field_operator_user() -> Any:
        return SimpleNamespace(
            id=uuid4(),
            role=UserRoleEnum.field_operator,
            is_active=True,
            email="field@test.local",
        )

    monkeypatch.setattr(ReportService, "generate", fake_generate)
    app.dependency_overrides[get_current_user] = field_operator_user
    try:
        response = await client.post(
            f"/api/v1/analytics/{uuid4()}/reports/generate",
            json={
                "irrigation_horizon_days": 7,
                "include_members": True,
                "include_history_charts": True,
            },
        )
        assert response.status_code == 403
    finally:
        app.dependency_overrides.pop(get_current_user, None)
