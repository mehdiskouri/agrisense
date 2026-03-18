"""Spreadsheet report generation service for farm analytics."""

from __future__ import annotations

import asyncio
import hashlib
import io
import json
import uuid
from datetime import UTC, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from redis.asyncio import Redis
from reportlab.lib.pagesizes import letter  # type: ignore[import-untyped]
from reportlab.pdfgen import canvas  # type: ignore[import-untyped]
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import get_settings
from app.schemas.reports import ReportJobCreateResponse, ReportJobStatusResponse, ReportRequest
from app.services.analytics_service import AnalyticsService
from app.services.farm_service import FarmService


class ReportService:
    JOB_TTL_SECONDS = 60 * 60 * 6
    FILE_TTL_SECONDS = 60 * 60 * 6
    SYNC_CACHE_TTL_SECONDS = 60 * 15
    HISTORY_CADENCE_MINUTES = 15

    def __init__(self, db: AsyncSession, redis_client: Redis | None = None):
        self.db = db
        self.redis_client = redis_client
        self.settings = get_settings()

    @staticmethod
    def _job_key(job_id: uuid.UUID) -> str:
        return f"reports:job:{job_id}"

    @staticmethod
    def _file_key(job_id: uuid.UUID) -> str:
        return f"reports:file:{job_id}"

    @staticmethod
    def _sync_cache_key(farm_id: uuid.UUID, request: ReportRequest, output_format: str) -> str:
        payload = request.model_dump(mode="json")
        encoded = json.dumps(payload, sort_keys=True, separators=(",", ":"))
        digest = hashlib.sha256(encoded.encode("utf-8")).hexdigest()
        return f"reports:sync:{output_format}:{farm_id}:{digest}"

    @staticmethod
    def _media_type(output_format: str) -> str:
        if output_format == "pdf":
            return "application/pdf"
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    @staticmethod
    def _filename(farm_id: uuid.UUID, output_format: str, at: datetime | None = None) -> str:
        ts = (at or datetime.now(UTC)).strftime("%Y%m%dT%H%M%SZ")
        extension = "pdf" if output_format == "pdf" else "xlsx"
        return f"agrisense-report-{farm_id}-{ts}.{extension}"

    @staticmethod
    def _validate_output_format(output_format: str) -> str:
        normalized = output_format.lower().strip()
        if normalized not in {"xlsx", "pdf"}:
            raise ValueError("unsupported output format")
        return normalized

    async def generate_report_artifact(
        self,
        farm_id: uuid.UUID,
        request: ReportRequest,
        *,
        output_format: str = "xlsx",
    ) -> tuple[bytes, str, str, bool]:
        output_format = self._validate_output_format(output_format)
        cache_key = self._sync_cache_key(farm_id, request, output_format)

        if self.redis_client is not None:
            cached_raw = await self.redis_client.get(cache_key)
            if cached_raw is not None:
                cached_bytes = (
                    cached_raw.encode("utf-8") if isinstance(cached_raw, str) else bytes(cached_raw)
                )
                return (
                    cached_bytes,
                    self._filename(farm_id, output_format),
                    self._media_type(output_format),
                    True,
                )

        if output_format == "pdf":
            content = await self._generate_pdf(farm_id, request)
        else:
            workbook = await self.generate(farm_id, request)
            content = workbook.getvalue()

        if self.redis_client is not None:
            await self.redis_client.setex(cache_key, self.SYNC_CACHE_TTL_SECONDS, content)

        return (
            content,
            self._filename(farm_id, output_format),
            self._media_type(output_format),
            False,
        )

    async def create_report_job(
        self,
        farm_id: uuid.UUID,
        request: ReportRequest,
        *,
        output_format: str = "xlsx",
    ) -> ReportJobCreateResponse:
        if self.redis_client is None:
            raise RuntimeError("redis unavailable for report async job")

        output_format = self._validate_output_format(output_format)

        await FarmService(self.db).get_farm(farm_id)
        job_id = uuid.uuid4()
        now = datetime.now(UTC)
        details_payload: dict[str, Any] = request.model_dump(mode="json")
        details_payload["output_format"] = output_format
        payload = {
            "job_id": str(job_id),
            "farm_id": str(farm_id),
            "status": "queued",
            "created_at": now.isoformat(),
            "updated_at": now.isoformat(),
            "completed_at": None,
            "error": None,
            "filename": None,
            "details": details_payload,
        }
        await self.redis_client.setex(
            self._job_key(job_id),
            self.JOB_TTL_SECONDS,
            json.dumps(payload),
        )
        return ReportJobCreateResponse(
            job_id=job_id,
            farm_id=farm_id,
            status="queued",
            created_at=now,
        )

    async def execute_report_job(
        self,
        job_id: uuid.UUID,
        farm_id: uuid.UUID,
        request: ReportRequest,
        *,
        output_format: str = "xlsx",
    ) -> ReportJobStatusResponse:
        if self.redis_client is None:
            raise RuntimeError("redis unavailable for report async job")

        output_format = self._validate_output_format(output_format)

        now = datetime.now(UTC)
        running_details: dict[str, Any] = request.model_dump(mode="json")
        running_details["output_format"] = output_format
        await self.redis_client.setex(
            self._job_key(job_id),
            self.JOB_TTL_SECONDS,
            json.dumps(
                {
                    "job_id": str(job_id),
                    "farm_id": str(farm_id),
                    "status": "running",
                    "created_at": now.isoformat(),
                    "updated_at": now.isoformat(),
                    "completed_at": None,
                    "error": None,
                    "filename": None,
                    "details": running_details,
                }
            ),
        )

        try:
            content, filename, _, _ = await self.generate_report_artifact(
                farm_id,
                request,
                output_format=output_format,
            )
            completed = datetime.now(UTC)
            await self.redis_client.setex(self._file_key(job_id), self.FILE_TTL_SECONDS, content)
            payload = {
                "job_id": str(job_id),
                "farm_id": str(farm_id),
                "status": "succeeded",
                "created_at": now.isoformat(),
                "updated_at": completed.isoformat(),
                "completed_at": completed.isoformat(),
                "error": None,
                "filename": filename,
                "details": {
                    "size_bytes": len(content),
                    "include_history_charts": request.include_history_charts,
                    "include_members": request.include_members,
                    "irrigation_horizon_days": request.irrigation_horizon_days,
                    "output_format": output_format,
                },
            }
            await self.redis_client.setex(
                self._job_key(job_id),
                self.JOB_TTL_SECONDS,
                json.dumps(payload),
            )
            return self._status_from_payload(payload)
        except Exception as exc:
            failed_at = datetime.now(UTC)
            failed_details: dict[str, Any] = request.model_dump(mode="json")
            failed_details["output_format"] = output_format
            payload = {
                "job_id": str(job_id),
                "farm_id": str(farm_id),
                "status": "failed",
                "created_at": now.isoformat(),
                "updated_at": failed_at.isoformat(),
                "completed_at": failed_at.isoformat(),
                "error": str(exc),
                "filename": None,
                "details": failed_details,
            }
            await self.redis_client.setex(
                self._job_key(job_id),
                self.JOB_TTL_SECONDS,
                json.dumps(payload),
            )
            return self._status_from_payload(payload)

    async def get_report_job_status(self, job_id: uuid.UUID) -> ReportJobStatusResponse:
        if self.redis_client is None:
            raise RuntimeError("redis unavailable for report async job")

        raw = await self.redis_client.get(self._job_key(job_id))
        if raw is None:
            raise LookupError(f"Report job {job_id} not found")
        payload = json.loads(raw)
        if not isinstance(payload, dict):
            raise LookupError(f"Report job {job_id} payload is invalid")
        return self._status_from_payload(payload)

    async def get_report_job_file(self, job_id: uuid.UUID) -> tuple[bytes, str, str]:
        if self.redis_client is None:
            raise RuntimeError("redis unavailable for report async job")

        status_payload = await self.get_report_job_status(job_id)
        if status_payload.status != "succeeded":
            raise ValueError(f"report job {job_id} is not ready")

        raw = await self.redis_client.get(self._file_key(job_id))
        if raw is None:
            raise LookupError(f"Report file for job {job_id} not found")
        content = raw.encode("utf-8") if isinstance(raw, str) else bytes(raw)
        details = status_payload.details if isinstance(status_payload.details, dict) else {}
        output_format = str(details.get("output_format") or "xlsx")
        output_format = self._validate_output_format(output_format)
        filename = status_payload.filename or f"agrisense-report-{job_id}.{output_format}"
        return content, filename, self._media_type(output_format)

    async def _generate_pdf(self, farm_id: uuid.UUID, request: ReportRequest) -> bytes:
        farm_service = FarmService(self.db)
        analytics = AnalyticsService(self.db, self.redis_client)

        farm = await farm_service.get_farm(farm_id)
        farm_status, irrigation, yield_forecast, nutrients, alerts = await asyncio.gather(
            analytics.get_farm_status(farm_id),
            analytics.get_irrigation_schedule(
                farm_id,
                horizon_days=request.irrigation_horizon_days,
            ),
            analytics.get_ensemble_yield_forecast(
                farm_id,
                include_members=request.include_members,
            ),
            analytics.get_nutrient_report(farm_id),
            analytics.get_active_alerts(farm_id),
        )

        buffer = io.BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=letter)
        height = float(letter[1])
        top = height - 50

        def new_page() -> float:
            pdf.showPage()
            return height - 50

        def write_line(cursor: float, text: str, *, bold: bool = False, indent: int = 0) -> float:
            if cursor < 60:
                cursor = new_page()
            pdf.setFont("Helvetica-Bold" if bold else "Helvetica", 10)
            pdf.drawString(50 + indent, cursor, text[:110])
            return cursor - 14

        y = top
        y = write_line(y, f"AgriSense Report: {farm.name}", bold=True)
        y = write_line(y, f"Farm ID: {farm.id}")
        y = write_line(y, f"Farm Type: {self._enum_str(farm.farm_type)}")
        y = write_line(y, f"Timezone: {farm.timezone}")
        y = write_line(y, f"Generated At: {datetime.now(UTC).isoformat()}")

        y = write_line(y - 6, "Farm Summary", bold=True)
        for zone in farm.zones:
            zone_summary = (
                f"- {zone.name} ({self._enum_str(zone.zone_type)}), "
                f"area={float(zone.area_m2):.2f} m2"
            )
            y = write_line(
                y,
                zone_summary,
                indent=8,
            )

        y = write_line(y - 6, "Irrigation Schedule", bold=True)
        for irrigation_item in irrigation.items[:20]:
            zone_token = irrigation_item.get("zone_id") or irrigation_item.get("zone")
            day_token = irrigation_item.get("day") or irrigation_item.get("date")
            should_irrigate = bool(
                irrigation_item.get("irrigate") or irrigation_item.get("should_irrigate") or False
            )
            priority = self._to_float(
                irrigation_item.get("priority") or irrigation_item.get("priority_score")
            )
            line = (
                f"- zone={zone_token}, day={day_token}, "
                f"irrigate={should_irrigate}, priority={priority:.2f}"
            )
            y = write_line(
                y,
                line,
                indent=8,
            )

        y = write_line(y - 6, "Yield Forecast", bold=True)
        for yield_item in yield_forecast.items[:20]:
            crop_bed = yield_item.crop_bed_id
            est = float(yield_item.yield_estimate_kg_m2)
            lower = float(yield_item.yield_lower)
            upper = float(yield_item.yield_upper)
            confidence = float(yield_item.confidence)
            line = (
                f"- crop_bed={crop_bed}, est={est:.3f}, "
                f"ci=[{lower:.3f}, {upper:.3f}], conf={confidence:.2f}"
            )
            y = write_line(
                y,
                line,
                indent=8,
            )

        y = write_line(y - 6, "Nutrient Status", bold=True)
        for nutrient_item in nutrients.items[:20]:
            zone_token = nutrient_item.get("zone_id")
            n_value = self._to_float(nutrient_item.get("nitrogen_deficit"))
            p_value = self._to_float(nutrient_item.get("phosphorus_deficit"))
            k_value = self._to_float(nutrient_item.get("potassium_deficit"))
            urgency = nutrient_item.get("urgency")
            line = (
                f"- zone={zone_token}, N={n_value:.2f}, "
                f"P={p_value:.2f}, K={k_value:.2f}, urgency={urgency}"
            )
            y = write_line(
                y,
                line,
                indent=8,
            )

        y = write_line(y - 6, "Active Alerts", bold=True)
        total_alerts = sum(len(zone_alerts.alerts) for zone_alerts in alerts.zones)
        y = write_line(y, f"Reported alert groups: {len(alerts.zones)}", indent=8)
        y = write_line(y, f"Zone status records: {len(farm_status.zones)}", indent=8)
        y = write_line(y, f"Derived status entries: {total_alerts}", indent=8)
        for zone_alerts in alerts.zones[:20]:
            for alert in zone_alerts.alerts[:5]:
                y = write_line(
                    y,
                    f"- zone={zone_alerts.zone_id}, src={alert.source}, sev={alert.severity}",
                    indent=8,
                )

        pdf.save()
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def _status_from_payload(payload: dict[str, Any]) -> ReportJobStatusResponse:
        completed_at_raw = payload.get("completed_at")
        completed_at = (
            datetime.fromisoformat(completed_at_raw) if isinstance(completed_at_raw, str) else None
        )
        details = payload.get("details")
        return ReportJobStatusResponse(
            job_id=uuid.UUID(str(payload["job_id"])),
            farm_id=uuid.UUID(str(payload["farm_id"])),
            status=str(payload.get("status") or "queued"),
            created_at=datetime.fromisoformat(str(payload["created_at"])),
            updated_at=datetime.fromisoformat(str(payload["updated_at"])),
            completed_at=completed_at,
            error=str(payload["error"]) if payload.get("error") is not None else None,
            filename=str(payload["filename"]) if payload.get("filename") is not None else None,
            details=details if isinstance(details, dict) else {},
        )

    async def generate(self, farm_id: uuid.UUID, request: ReportRequest) -> io.BytesIO:
        farm_service = FarmService(self.db)
        analytics = AnalyticsService(self.db, self.redis_client)

        farm = await farm_service.get_farm(farm_id)

        (
            farm_status,
            irrigation,
            yield_forecast,
            nutrients,
            alerts,
        ) = await asyncio.gather(
            analytics.get_farm_status(farm_id),
            analytics.get_irrigation_schedule(
                farm_id,
                horizon_days=request.irrigation_horizon_days,
            ),
            analytics.get_ensemble_yield_forecast(
                farm_id,
                include_members=request.include_members,
            ),
            analytics.get_nutrient_report(farm_id),
            analytics.get_active_alerts(farm_id),
        )

        graph_state: dict[str, Any] | None = None
        if request.include_history_charts:
            graph_state = await farm_service.get_graph(farm_id)

        workbook = Workbook()
        default_sheet = workbook.active
        if default_sheet is not None:
            workbook.remove(default_sheet)

        self._build_farm_summary_sheet(workbook, farm, farm_status)
        self._build_irrigation_sheet(workbook, irrigation)
        self._build_yield_sheet(workbook, yield_forecast, request.include_members)
        self._build_nutrient_sheet(workbook, nutrients)
        self._build_alerts_sheet(workbook, alerts)
        if request.include_history_charts:
            self._build_history_sheet(workbook, graph_state)

        for sheet in workbook.worksheets:
            self._auto_fit_columns(sheet)

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    def _build_farm_summary_sheet(self, workbook: Workbook, farm: Any, farm_status: Any) -> None:
        sheet = workbook.create_sheet("Farm Summary")
        title = f"AgriSense Report - {farm.name}"
        sheet.merge_cells("A1:F1")
        sheet["A1"] = title
        sheet["A1"].font = Font(size=14, bold=True, color="FFFFFF")
        sheet["A1"].fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
        sheet["A1"].alignment = Alignment(horizontal="center")

        sheet["A3"] = "Farm ID"
        sheet["B3"] = str(farm.id)
        sheet["A4"] = "Farm Type"
        sheet["B4"] = self._enum_str(farm.farm_type)
        sheet["A5"] = "Timezone"
        sheet["B5"] = str(farm.timezone)
        sheet["A6"] = "Generated At"
        sheet["B6"] = datetime.now(UTC).isoformat()

        headers = [
            "Zone ID",
            "Zone Name",
            "Zone Type",
            "Area (m2)",
            "Soil Type",
            "Layer Status",
        ]
        start_row = 8
        self._write_header_row(sheet, start_row, headers, fill_hex="2D6A4F")

        status_index: dict[str, dict[str, Any]] = {
            str(item.zone_id): item.status for item in farm_status.zones
        }

        row = start_row + 1
        for zone in farm.zones:
            layer_status = status_index.get(str(zone.id), {})
            compact_status = ", ".join(sorted(layer_status.keys())) or "n/a"
            sheet.append(
                [
                    str(zone.id),
                    str(zone.name),
                    self._enum_str(zone.zone_type),
                    float(zone.area_m2),
                    str(zone.soil_type),
                    compact_status,
                ]
            )
            row += 1

    def _build_irrigation_sheet(self, workbook: Workbook, irrigation: Any) -> None:
        sheet = workbook.create_sheet("Irrigation Schedule")
        headers = [
            "Zone",
            "Day",
            "Irrigate",
            "Volume (L)",
            "Priority",
            "Projected Moisture",
            "Trigger Reason",
        ]
        self._write_header_row(sheet, 1, headers, fill_hex="1D4ED8")

        for item in irrigation.items:
            zone_token = item.get("zone_id") or item.get("zone") or "n/a"
            day = item.get("day") or item.get("date") or "n/a"
            irrigate = bool(item.get("irrigate") or item.get("should_irrigate") or False)
            volume = self._to_float(item.get("volume_liters") or item.get("volume_l"))
            priority = self._to_float(item.get("priority") or item.get("priority_score"))
            projected = self._to_float(item.get("projected_moisture") or item.get("moisture_after"))
            reason = str(item.get("trigger_reason") or item.get("reason") or "")
            sheet.append([str(zone_token), str(day), irrigate, volume, priority, projected, reason])

        irrigate_col = "C"
        first_data = 2
        last_data = max(2, sheet.max_row)
        green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        gray_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        sheet.conditional_formatting.add(
            f"{irrigate_col}{first_data}:{irrigate_col}{last_data}",
            self._cell_is_rule(operator="equal", formula=["TRUE"], fill=green_fill),
        )
        sheet.conditional_formatting.add(
            f"{irrigate_col}{first_data}:{irrigate_col}{last_data}",
            self._cell_is_rule(operator="equal", formula=["FALSE"], fill=gray_fill),
        )
        sheet.conditional_formatting.add(
            f"E{first_data}:E{last_data}",
            self._color_scale_rule(
                start_type="num",
                start_value=0,
                start_color="DCFCE7",
                mid_type="num",
                mid_value=0.5,
                mid_color="FEF9C3",
                end_type="num",
                end_value=1,
                end_color="FCA5A5",
            ),
        )

    def _build_yield_sheet(
        self, workbook: Workbook, yield_forecast: Any, include_members: bool
    ) -> None:
        sheet = workbook.create_sheet("Yield Forecast")
        headers = [
            "Crop Bed",
            "Estimate",
            "Lower CI",
            "Upper CI",
            "Confidence",
            "Ks",
            "Kn",
            "Kl",
            "Kw",
            "Model Layer",
        ]
        self._write_header_row(sheet, 1, headers, fill_hex="166534")

        row = 2
        for item in yield_forecast.items:
            stress = item.stress_factors
            sheet.append(
                [
                    str(item.crop_bed_id),
                    float(item.yield_estimate_kg_m2),
                    float(item.yield_lower),
                    float(item.yield_upper),
                    float(item.confidence),
                    self._to_float(stress.get("Ks")),
                    self._to_float(stress.get("Kn")),
                    self._to_float(stress.get("Kl")),
                    self._to_float(stress.get("Kw")),
                    str(item.model_layer),
                ]
            )
            row += 1

        conf_fill_hi = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        conf_fill_mid = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
        conf_fill_low = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        end_row = max(2, sheet.max_row)
        sheet.conditional_formatting.add(
            f"E2:E{end_row}",
            self._cell_is_rule(operator="greaterThanOrEqual", formula=["0.8"], fill=conf_fill_hi),
        )
        sheet.conditional_formatting.add(
            f"E2:E{end_row}",
            self._cell_is_rule(operator="between", formula=["0.5", "0.7999"], fill=conf_fill_mid),
        )
        sheet.conditional_formatting.add(
            f"E2:E{end_row}",
            self._cell_is_rule(operator="lessThan", formula=["0.5"], fill=conf_fill_low),
        )

        if sheet.max_row >= 2:
            chart = LineChart()
            chart.title = "Yield Estimate With Confidence Interval"
            chart.y_axis.title = "kg/m2"
            chart.x_axis.title = "Crop Bed"
            data_ref = Reference(sheet, min_col=2, min_row=1, max_col=4, max_row=sheet.max_row)
            cat_ref = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cat_ref)
            chart.height = 8
            chart.width = 20
            sheet.add_chart(chart, "L2")

        summary_row = sheet.max_row + 3
        sheet.cell(row=summary_row, column=1, value="Ensemble Weights Summary")
        sheet.cell(row=summary_row, column=1).font = Font(bold=True)
        self._write_header_row(sheet, summary_row + 1, ["Model", "Weight"], fill_hex="047857")
        weight_rows_start = summary_row + 2
        ensemble_weights = getattr(yield_forecast, "ensemble_weights", None)
        if isinstance(ensemble_weights, dict):
            for model_name, weight in ensemble_weights.items():
                sheet.append([str(model_name), self._to_float(weight)])
        else:
            sheet.append(["n/a", 0.0])

        if include_members:
            member_row = max(sheet.max_row + 3, weight_rows_start + 2)
            self._write_header_row(
                sheet,
                member_row,
                ["Crop Bed", "Member Model", "Estimate", "Lower", "Upper", "Weight"],
                fill_hex="065F46",
            )
            member_row += 1
            for item in yield_forecast.items:
                for member in item.ensemble_members:
                    sheet.append(
                        [
                            str(item.crop_bed_id),
                            str(member.model_name),
                            float(member.yield_estimate),
                            float(member.lower),
                            float(member.upper),
                            float(member.weight),
                        ]
                    )

    def _build_nutrient_sheet(self, workbook: Workbook, nutrients: Any) -> None:
        sheet = workbook.create_sheet("Nutrient Status")
        headers = [
            "Zone",
            "N Deficit",
            "P Deficit",
            "K Deficit",
            "Severity Score",
            "Urgency",
            "Suggested Amendment",
            "Visual Confirmed",
        ]
        self._write_header_row(sheet, 1, headers, fill_hex="C2410C")

        for item in nutrients.items:
            n_def = self._to_float(item.get("nitrogen_deficit"))
            p_def = self._to_float(item.get("phosphorus_deficit"))
            k_def = self._to_float(item.get("potassium_deficit"))
            severity = round((n_def + p_def + k_def) / 3.0, 4)
            urgency = str(item.get("urgency") or "low")
            amendment = str(item.get("suggested_amendment") or item.get("recommendation") or "")
            visual = bool(item.get("visual_confirmed") or False)
            sheet.append(
                [
                    str(item.get("zone_id") or "n/a"),
                    n_def,
                    p_def,
                    k_def,
                    severity,
                    urgency,
                    amendment,
                    visual,
                ]
            )

        urgency_col = "F"
        end_row = max(2, sheet.max_row)
        sheet.conditional_formatting.add(
            f"{urgency_col}2:{urgency_col}{end_row}",
            self._cell_is_rule(
                operator="equal",
                formula=['"critical"'],
                fill=PatternFill("solid", "B91C1C", "B91C1C"),
            ),
        )
        sheet.conditional_formatting.add(
            f"{urgency_col}2:{urgency_col}{end_row}",
            self._cell_is_rule(
                operator="equal", formula=['"high"'], fill=PatternFill("solid", "EA580C", "EA580C")
            ),
        )
        sheet.conditional_formatting.add(
            f"{urgency_col}2:{urgency_col}{end_row}",
            self._cell_is_rule(
                operator="equal",
                formula=['"medium"'],
                fill=PatternFill("solid", "F59E0B", "F59E0B"),
            ),
        )
        sheet.conditional_formatting.add(
            f"{urgency_col}2:{urgency_col}{end_row}",
            self._cell_is_rule(
                operator="equal", formula=['"low"'], fill=PatternFill("solid", "16A34A", "16A34A")
            ),
        )
        sheet.conditional_formatting.add(
            f"E2:E{end_row}",
            self._color_scale_rule(
                start_type="num",
                start_value=0,
                start_color="16A34A",
                mid_type="num",
                mid_value=0.5,
                mid_color="F59E0B",
                end_type="num",
                end_value=1,
                end_color="B91C1C",
            ),
        )

    def _build_alerts_sheet(self, workbook: Workbook, alerts: Any) -> None:
        sheet = workbook.create_sheet("Active Alerts")
        headers = [
            "Zone ID",
            "Source",
            "Severity",
            "Layer",
            "Anomaly Type",
            "Urgency",
            "Suggestion",
        ]
        self._write_header_row(sheet, 1, headers, fill_hex="991B1B")

        for zone_alerts in alerts.zones:
            zone_id = str(zone_alerts.zone_id) if zone_alerts.zone_id is not None else "n/a"
            for alert in zone_alerts.alerts:
                payload = alert.payload
                sheet.append(
                    [
                        zone_id,
                        str(alert.source),
                        str(alert.severity),
                        str(payload.get("layer") or ""),
                        str(payload.get("anomaly_type") or ""),
                        str(payload.get("urgency") or ""),
                        str(payload.get("suggestion") or payload.get("recommended_action") or ""),
                    ]
                )

        end_row = max(2, sheet.max_row)
        sheet.conditional_formatting.add(
            f"C2:C{end_row}",
            self._cell_is_rule(
                operator="equal",
                formula=['"critical"'],
                fill=PatternFill("solid", "7F1D1D", "7F1D1D"),
                font=Font(color="FFFFFF"),
            ),
        )
        sheet.conditional_formatting.add(
            f"C2:C{end_row}",
            self._cell_is_rule(
                operator="equal",
                formula=['"warning"'],
                fill=PatternFill("solid", "F59E0B", "F59E0B"),
            ),
        )
        sheet.conditional_formatting.add(
            f"C2:C{end_row}",
            self._cell_is_rule(
                operator="equal", formula=['"info"'], fill=PatternFill("solid", "1D4ED8", "1D4ED8")
            ),
        )
        sheet.auto_filter.ref = f"A1:G{end_row}"

    def _build_history_sheet(self, workbook: Workbook, graph_state: dict[str, Any] | None) -> None:
        sheet = workbook.create_sheet("Feature History")

        if graph_state is None:
            return

        layers_raw = graph_state.get("layers")
        if not isinstance(layers_raw, dict):
            return

        row = 1
        max_points = max(1, int(self.settings.report_max_history_points))
        layer_header_colors = ["334155", "0F766E", "1D4ED8", "7C3AED", "9A3412", "166534"]

        for layer_index, (layer_name, layer_payload) in enumerate(sorted(layers_raw.items())):
            if not isinstance(layer_payload, dict):
                continue

            history = layer_payload.get("feature_history")
            if not isinstance(history, list) or not history:
                continue

            first_vertex = history[0]
            if not isinstance(first_vertex, list) or not first_vertex:
                continue

            feature_dim = len(first_vertex)
            buffer_size = len(first_vertex[0]) if isinstance(first_vertex[0], list) else 0
            if feature_dim == 0 or buffer_size == 0:
                continue

            history_length = self._to_int(layer_payload.get("history_length"), fallback=0)
            if history_length <= 0:
                continue
            history_head = self._to_int(layer_payload.get("history_head"), fallback=1)

            ordered = self._history_indices(history_length, buffer_size, history_head)
            if not ordered:
                continue
            ordered = ordered[-max_points:]

            header_fill = layer_header_colors[layer_index % len(layer_header_colors)]
            self._write_header_row(
                sheet,
                row,
                [
                    f"Layer: {layer_name}",
                    "Time Index (+min)",
                    "Feature 1",
                    "Feature 2",
                    "Feature 3",
                    "Feature 4",
                    "Feature 5",
                ],
                fill_hex=header_fill,
            )

            section_start = row + 1
            row += 1
            for point_index, slot in enumerate(ordered, start=1):
                features: list[float] = []
                for f_idx in range(feature_dim):
                    feature_series = first_vertex[f_idx]
                    if isinstance(feature_series, list) and slot < len(feature_series):
                        features.append(self._to_float(feature_series[slot]))
                    else:
                        features.append(0.0)

                minute_index = (point_index - 1) * self.HISTORY_CADENCE_MINUTES
                values: list[Any] = [str(layer_name), minute_index]
                values.extend(features[:5])
                while len(values) < 7:
                    values.append(None)
                sheet.append(values)
                row += 1

            if row - section_start >= 2:
                chart = LineChart()
                chart.title = f"{layer_name} feature history"
                chart.y_axis.title = "Value"
                chart.x_axis.title = "Time (+min)"
                max_col = min(2 + feature_dim, 7)
                data_ref = Reference(
                    sheet,
                    min_col=3,
                    min_row=section_start - 1,
                    max_col=max_col,
                    max_row=row - 1,
                )
                cat_ref = Reference(sheet, min_col=2, min_row=section_start, max_row=row - 1)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cat_ref)
                chart.height = 6
                chart.width = 12
                sheet.add_chart(chart, f"I{section_start}")

            row += 2

    @staticmethod
    def _write_header_row(
        sheet: Worksheet,
        row_index: int,
        headers: list[str],
        *,
        fill_hex: str,
    ) -> None:
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row_index, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

    @staticmethod
    def _auto_fit_columns(sheet: Worksheet) -> None:
        for index, column in enumerate(sheet.columns, start=1):
            values = [str(cell.value) for cell in column if cell.value is not None]
            max_len = max((len(value) for value in values), default=0)
            adjusted = min(max(12, max_len + 2), 60)
            letter = get_column_letter(index)
            sheet.column_dimensions[letter].width = adjusted

    @staticmethod
    def _to_float(value: Any) -> float:
        if isinstance(value, (int, float)):
            return float(value)
        try:
            return float(str(value))
        except (TypeError, ValueError):
            return 0.0

    @staticmethod
    def _to_int(value: Any, fallback: int) -> int:
        if isinstance(value, int):
            return value
        try:
            return int(str(value))
        except (TypeError, ValueError):
            return fallback

    @staticmethod
    def _enum_str(value: Any) -> str:
        enum_value = getattr(value, "value", None)
        if isinstance(enum_value, str):
            return enum_value
        return str(value)

    @staticmethod
    def _history_indices(history_length: int, buffer_size: int, history_head: int) -> list[int]:
        if history_length <= 0 or buffer_size <= 0:
            return []

        if history_length < buffer_size:
            return [index for index in range(history_length)]

        # history_head is 1-indexed next-write slot; convert to 0-index oldest slot.
        start = max(0, history_head - 1)
        indices: list[int] = []
        for offset in range(buffer_size):
            indices.append((start + offset) % buffer_size)
        return indices

    @staticmethod
    def _cell_is_rule(**kwargs: Any) -> Any:
        return CellIsRule(**kwargs)  # type: ignore[no-untyped-call]

    @staticmethod
    def _color_scale_rule(**kwargs: Any) -> Any:
        return ColorScaleRule(**kwargs)  # type: ignore[no-untyped-call]
